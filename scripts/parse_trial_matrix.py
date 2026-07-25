import openpyxl

wb = openpyxl.load_workbook('/Users/daniyarovaruslanovna/Downloads/набор дк 50 (4).xlsx', data_only=True)
sheet = wb.active

# Rows 3 to 12 contain the trial lens grid A1-V10
headers = [sheet.cell(row=2, column=c).value for c in range(2, 16)]
print("Columns headers (FK groups):", headers)

for r in range(3, 13):
    print(f"\n--- GRID ROW {r-2} ---")
    for c in range(2, 16):
        col_header = sheet.cell(row=2, column=c).value
        cell_val = sheet.cell(row=r, column=c).value
        if cell_val:
            print(f"Col {col_header} (r={r}, c={c}): {str(cell_val).strip()}")

