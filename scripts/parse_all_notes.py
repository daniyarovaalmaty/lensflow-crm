import openpyxl

wb = openpyxl.load_workbook('/Users/daniyarovaruslanovna/Downloads/MediLens Калькулятор й.xlsx', data_only=True)
sheet = wb.active

print("=== MediLens Calculator Notes & Rules ===")
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=r, column=c).value
        if v:
            print(f"Row {r}, Col {c}: {str(v).strip()}")
