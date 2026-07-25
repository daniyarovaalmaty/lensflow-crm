import openpyxl

wb = openpyxl.load_workbook('/Users/daniyarovaruslanovna/Downloads/MediLens Калькулятор й.xlsx', data_only=False)
wb_val = openpyxl.load_workbook('/Users/daniyarovaruslanovna/Downloads/MediLens Калькулятор й.xlsx', data_only=True)

sheet = wb.active
sheet_val = wb_val.active

print(f"Sheet name: {sheet.title}, Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        cell_f = sheet.cell(row=r, column=c).value
        cell_v = sheet_val.cell(row=r, column=c).value
        if cell_f is not None:
            col = openpyxl.utils.get_column_letter(c)
            print(f"{col}{r}: FORMULA='{cell_f}' | EVAL='{cell_v}'")
