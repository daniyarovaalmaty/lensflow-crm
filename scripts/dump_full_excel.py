import openpyxl

def dump_excel(filepath):
    print(f"\n==========================================")
    print(f"FULL DUMP: {filepath}")
    print(f"==========================================")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    wb_val = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_val = wb_val[sheet_name]
        print(f"\n--- SHEET: {sheet_name} ---")
        for r in range(1, sheet.max_row + 1):
            row_str = []
            for c in range(1, sheet.max_column + 1):
                formula = sheet.cell(row=r, column=c).value
                val = sheet_val.cell(row=r, column=c).value
                if formula is not None:
                    col_let = openpyxl.utils.get_column_letter(c)
                    if str(formula).startswith('='):
                        row_str.append(f"{col_let}{r}={formula} [eval={val}]")
                    else:
                        row_str.append(f"{col_let}{r}={formula}")
            if row_str:
                print(" | ".join(row_str))

dump_excel('/Users/daniyarovaruslanovna/Downloads/MediLens Калькулятор й.xlsx')
dump_excel('/Users/daniyarovaruslanovna/Downloads/набор дк 50 (4).xlsx')
