import openpyxl

wb = openpyxl.load_workbook('/Users/daniyarovaruslanovna/Downloads/набор дк 50 (4).xlsx', data_only=True)
sheet = wb.active

print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} cols")

for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
    non_empty = [(openpyxl.utils.get_column_letter(c), val) for c, val in enumerate(row_vals, 1) if val is not None]
    if non_empty:
        print(f"Row {r:2d}: " + " | ".join([f"{col}={str(val).strip()}" for col, val in non_empty]))

