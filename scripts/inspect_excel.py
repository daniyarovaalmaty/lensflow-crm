import openpyxl

def inspect_file(filepath):
    print(f"=== Inspecting {filepath} ===")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} (rows: {sheet.max_row}, cols: {sheet.max_column}) ---")
        # Print non-empty cells in the first 30 rows
        for r in range(1, min(sheet.max_row + 1, 40)):
            row_vals = []
            for c in range(1, min(sheet.max_column + 1, 20)):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    row_vals.append(f"[{openpyxl.utils.get_column_letter(c)}{r}] {val}")
            if row_vals:
                print(" | ".join(row_vals[:10]))

inspect_file('/Users/daniyarovaruslanovna/Downloads/MediLens Калькулятор й.xlsx')
inspect_file('/Users/daniyarovaruslanovna/Downloads/набор дк 50 (4).xlsx')
